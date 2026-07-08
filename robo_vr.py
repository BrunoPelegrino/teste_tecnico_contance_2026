#!/usr/bin/env python3
"""
robo_vr.py
----------
Automação de cálculo e geração do relatório mensal de Vale-Refeição (VR)
para a área de RH da Constance Calçados.

Autor: Bruno (Analista de Automação - teste técnico Constance)

Uso:
    python robo_vr.py
    python robo_vr.py --entrada colaboradores.xlsx --saida relatorio_vr_072026.xlsx
    python robo_vr.py --data-ref 2026-07-01

Ver README.md para instruções completas e DECISOES.md para justificativas
técnicas das decisões de implementação.
"""

from __future__ import annotations

import argparse
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# CONSTANTES / REGRAS DE NEGÓCIO (seção 5 do teste técnico)
# =============================================================================

VALOR_VR_LOGISTICA = 35.00
VALOR_VR_PADRAO = 30.00
SETOR_LOGISTICA = "logística"  # comparado normalizado (lower + strip)

PERC_DESCONTO_COLABORADOR = 0.20
DIAS_EXPERIENCIA_LIMITE = 90
LIMITE_ABSENTEISMO_PCT = 10.0

COLUNAS_OBRIGATORIAS = [
    "matricula",
    "nome",
    "setor",
    "dias_trabalhados",
    "dias_uteis_mes",
    "faltas",
    "atestados",
    "admissao",
]

# Setores observados no cadastro da empresa. Usado apenas para alertar sobre
# possíveis erros de digitação; um setor fora desta lista NÃO é bloqueado
# (recebe o valor de VR padrão), apenas sinalizado como aviso.
SETORES_CONHECIDOS = {"logística", "rh", "ti", "financeiro", "comercial"}

# Paleta de formatação (seção 6.2 do teste técnico)
COR_HEADER_BG = "8C3A46"      # bordô, alinhado à identidade visual Constance
COR_HEADER_FONT = "FFFFFF"
COR_ZEBRA_PAR = "F5EDEC"
COR_ZEBRA_IMPAR = "FFFFFF"
COR_ALERTA_BG = "F8C9C4"      # vermelho claro

FORMATO_MOEDA = 'R$ #,##0.00'

# Pausa entre mensagens de progresso, para dar sensação de processamento real
# ao usuário do RH (em vez das mensagens aparecerem todas instantaneamente).
INTERVALO_PROGRESSO_SEGUNDOS = 1.0


# =============================================================================
# EXCEÇÕES CUSTOMIZADAS
# =============================================================================

class RoboVRError(Exception):
    """Exceção-base para erros conhecidos e tratados do robô."""


class ArquivoNaoEncontradoError(RoboVRError):
    pass


class EstruturaInvalidaError(RoboVRError):
    pass


class PlanilhaVaziaError(RoboVRError):
    pass


# =============================================================================
# MODELOS DE DADOS
# =============================================================================

@dataclass
class Colaborador:
    """Representa uma linha bruta (já com tipos validados) do arquivo de entrada."""
    linha_excel: int
    matricula: str
    nome: str
    setor: str
    dias_trabalhados: int
    dias_uteis_mes: int
    faltas: int
    atestados: int
    admissao: date


@dataclass
class ResultadoVR:
    """Representa uma linha já calculada, pronta para ir ao relatório de saída."""
    matricula: str
    nome: str
    setor: str
    dias_a_pagar: int
    valor_bruto_vr: float
    desconto_colaborador: float
    valor_liquido_empresa: float
    elegibilidade: str          # "Experiência (50%)" ou "Integral (100%)"
    absenteismo_pct: float
    status: str                 # "OK" ou "ALERTA"
    avisos: list[str] = field(default_factory=list)


@dataclass
class Estatisticas:
    total_colaboradores: int = 0
    total_pagar_empresa: float = 0.0
    total_alerta_absenteismo: int = 0
    total_em_experiencia: int = 0
    gerado_em: datetime = field(default_factory=datetime.now)


# =============================================================================
# CAMADA DE LEITURA
# =============================================================================

def caminho_base_execucao() -> Path:
    """
    Retorna o diretório onde o executável/script está rodando.
    Necessário porque, quando empacotado com PyInstaller (--onefile),
    sys.executable aponta para o .exe, não para o script Python original.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def localizar_arquivo_entrada(nome_arquivo: str) -> Path:
    """
    Procura o arquivo de entrada em dois lugares, nesta ordem:
    1) caminho informado, relativo ao diretório de trabalho atual (uso comum via CLI);
    2) mesma pasta do executável/script (requisito do .exe: seção 6.3 do teste técnico).
    """
    caminho_relativo_cwd = Path(nome_arquivo)
    if caminho_relativo_cwd.exists():
        return caminho_relativo_cwd

    caminho_junto_ao_programa = caminho_base_execucao() / nome_arquivo
    if caminho_junto_ao_programa.exists():
        return caminho_junto_ao_programa

    raise ArquivoNaoEncontradoError(
        f"Arquivo '{nome_arquivo}' não encontrado no diretório atual nem em "
        f"'{caminho_base_execucao()}'. Verifique se o arquivo está na mesma pasta do programa."
    )


def carregar_workbook(caminho: Path) -> Worksheet:
    """Abre o arquivo de entrada e retorna a planilha ativa."""
    try:
        wb = load_workbook(caminho, data_only=True)
    except Exception as exc:
        raise EstruturaInvalidaError(
            f"Não foi possível abrir '{caminho.name}'. O arquivo pode estar corrompido, "
            f"protegido por senha ou não ser um .xlsx válido. Detalhe técnico: {exc}"
        ) from exc
    return wb.active


def validar_estrutura_colunas(ws: Worksheet) -> dict[str, int]:
    """
    Lê o cabeçalho (linha 1) e retorna um mapa {nome_coluna: indice_coluna}.
    Isso permite ler por NOME, e não por posição fixa, tornando o robô
    resiliente a reordenação de colunas pelo RH em meses futuros.
    """
    primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if primeira_linha is None:
        raise PlanilhaVaziaError("A planilha de entrada está vazia (sem cabeçalho).")

    mapa_colunas = {}
    for idx, valor in enumerate(primeira_linha, start=1):
        if valor is not None:
            mapa_colunas[str(valor).strip().lower()] = idx

    faltantes = [c for c in COLUNAS_OBRIGATORIAS if c not in mapa_colunas]
    if faltantes:
        raise EstruturaInvalidaError(
            f"Coluna(s) obrigatória(s) ausente(s) no arquivo de entrada: {', '.join(faltantes)}. "
            f"Colunas esperadas: {', '.join(COLUNAS_OBRIGATORIAS)}."
        )
    return mapa_colunas


def _parse_data_admissao(valor: object, linha: int) -> date:
    """Converte a célula de admissão em date, aceitando datetime ou string ISO/BR."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        valor = valor.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(valor, fmt).date()
            except ValueError:
                continue
    raise EstruturaInvalidaError(
        f"Linha {linha}: data de admissão em formato inesperado ('{valor}'). "
        f"Use AAAA-MM-DD ou DD/MM/AAAA."
    )


def _parse_inteiro_nao_negativo(valor: object, nome_campo: str, linha: int) -> int:
    try:
        numero = int(valor)
    except (TypeError, ValueError):
        raise EstruturaInvalidaError(
            f"Linha {linha}: campo '{nome_campo}' inválido (valor: '{valor}'), esperado número inteiro."
        )
    if numero < 0:
        raise EstruturaInvalidaError(
            f"Linha {linha}: campo '{nome_campo}' não pode ser negativo (valor: {numero})."
        )
    return numero


def ler_colaboradores(ws: Worksheet, mapa_colunas: dict[str, int]) -> list[Colaborador]:
    """Lê todas as linhas de dados e as converte em objetos Colaborador validados."""
    colaboradores: list[Colaborador] = []

    for linha_idx, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if linha is None or all(v is None for v in linha):
            continue  # ignora linhas totalmente vazias (comum no fim de planilhas exportadas)

        def campo(nome_coluna: str) -> object:
            return linha[mapa_colunas[nome_coluna] - 1]

        matricula_val = campo("matricula")
        nome_val = campo("nome")

        if matricula_val is None or nome_val is None:
            raise EstruturaInvalidaError(
                f"Linha {linha_idx}: 'matricula' e 'nome' são obrigatórios e não podem estar vazios."
            )

        colaborador = Colaborador(
            linha_excel=linha_idx,
            matricula=str(matricula_val).strip(),
            nome=str(nome_val).strip(),
            setor=str(campo("setor") or "").strip(),
            dias_trabalhados=_parse_inteiro_nao_negativo(campo("dias_trabalhados"), "dias_trabalhados", linha_idx),
            dias_uteis_mes=_parse_inteiro_nao_negativo(campo("dias_uteis_mes"), "dias_uteis_mes", linha_idx),
            faltas=_parse_inteiro_nao_negativo(campo("faltas"), "faltas", linha_idx),
            atestados=_parse_inteiro_nao_negativo(campo("atestados"), "atestados", linha_idx),
            admissao=_parse_data_admissao(campo("admissao"), linha_idx),
        )
        colaboradores.append(colaborador)

    if not colaboradores:
        raise PlanilhaVaziaError("Nenhum colaborador encontrado no arquivo de entrada.")

    return colaboradores


# =============================================================================
# CAMADA DE VALIDAÇÃO (avisos que não interrompem o processamento)
# =============================================================================

def validar_colaborador(c: Colaborador, data_referencia: date) -> list[str]:
    """Retorna avisos de qualidade de dado (não fatais) para uma linha."""
    avisos: list[str] = []

    if c.dias_uteis_mes == 0:
        avisos.append("dias_uteis_mes igual a 0 (absenteísmo não pôde ser calculado).")

    if c.dias_trabalhados + c.faltas > c.dias_uteis_mes and c.dias_uteis_mes > 0:
        avisos.append(
            f"dias_trabalhados ({c.dias_trabalhados}) + faltas ({c.faltas}) "
            f"excede dias_uteis_mes ({c.dias_uteis_mes}) — verificar lançamento."
        )

    if c.admissao > data_referencia:
        avisos.append(
            f"data de admissão ({c.admissao.isoformat()}) é posterior à data de referência "
            f"({data_referencia.isoformat()}) — verificar se é um cadastro futuro válido."
        )

    if not c.setor:
        avisos.append("setor não informado — aplicado valor de VR padrão (demais setores).")
    elif c.setor.strip().lower() not in SETORES_CONHECIDOS:
        avisos.append(
            f"setor '{c.setor}' não reconhecido na lista de setores cadastrados "
            f"({', '.join(sorted(SETORES_CONHECIDOS))}) — aplicado valor de VR padrão. "
            f"Verifique se não é erro de digitação."
        )

    return avisos


# =============================================================================
# CAMADA DE REGRAS DE NEGÓCIO (funções puras, testáveis isoladamente)
# =============================================================================

def calcular_elegibilidade(c: Colaborador, data_referencia: date) -> tuple[float, str]:
    """Retorna (percentual_elegibilidade, rótulo) conforme regra 5.1."""
    dias_desde_admissao = (data_referencia - c.admissao).days
    if dias_desde_admissao < DIAS_EXPERIENCIA_LIMITE:
        return 0.5, "Experiência (50%)"
    return 1.0, "Integral (100%)"


def calcular_dias_a_pagar(c: Colaborador) -> int:
    """Regra 5.2: dias_trabalhados + atestados (faltas já descontadas)."""
    return c.dias_trabalhados + c.atestados


def valor_dia_por_setor(setor: str) -> float:
    """Regra 5.3: valor por dia conforme setor."""
    if setor.strip().lower() == SETOR_LOGISTICA:
        return VALOR_VR_LOGISTICA
    return VALOR_VR_PADRAO


def calcular_valor_vr(c: Colaborador, dias_a_pagar: int, percentual_elegibilidade: float) -> tuple[float, float, float]:
    """Regra 5.3 e 5.4: retorna (valor_bruto, desconto_colaborador, valor_liquido_empresa)."""
    valor_dia = valor_dia_por_setor(c.setor)
    valor_bruto = dias_a_pagar * valor_dia * percentual_elegibilidade
    desconto = valor_bruto * PERC_DESCONTO_COLABORADOR
    valor_liquido = valor_bruto - desconto
    return round(valor_bruto, 2), round(desconto, 2), round(valor_liquido, 2)


def calcular_absenteismo(c: Colaborador) -> tuple[float, bool]:
    """Regra 5.5: retorna (percentual, em_alerta). Protegido contra divisão por zero."""
    if c.dias_uteis_mes == 0:
        return 0.0, False
    percentual = (c.faltas / c.dias_uteis_mes) * 100
    em_alerta = percentual > LIMITE_ABSENTEISMO_PCT
    return round(percentual, 2), em_alerta


# =============================================================================
# CAMADA DE AGREGAÇÃO
# =============================================================================

def processar_colaboradores(
    colaboradores: list[Colaborador], data_referencia: date
) -> tuple[list[ResultadoVR], Estatisticas]:
    resultados: list[ResultadoVR] = []
    stats = Estatisticas()

    for c in colaboradores:
        avisos = validar_colaborador(c, data_referencia)
        for aviso in avisos:
            print(f"  [AVISO] Linha {c.linha_excel} ({c.nome}): {aviso}")

        percentual_elig, rotulo_elig = calcular_elegibilidade(c, data_referencia)
        dias_a_pagar = calcular_dias_a_pagar(c)
        valor_bruto, desconto, valor_liquido = calcular_valor_vr(c, dias_a_pagar, percentual_elig)
        absenteismo_pct, em_alerta = calcular_absenteismo(c)

        resultado = ResultadoVR(
            matricula=c.matricula,
            nome=c.nome,
            setor=c.setor,
            dias_a_pagar=dias_a_pagar,
            valor_bruto_vr=valor_bruto,
            desconto_colaborador=desconto,
            valor_liquido_empresa=valor_liquido,
            elegibilidade=rotulo_elig,
            absenteismo_pct=absenteismo_pct,
            status="ALERTA" if em_alerta else "OK",
            avisos=avisos,
        )
        resultados.append(resultado)

        stats.total_colaboradores += 1
        stats.total_pagar_empresa += valor_liquido
        if em_alerta:
            stats.total_alerta_absenteismo += 1
        if percentual_elig < 1.0:
            stats.total_em_experiencia += 1

    stats.total_pagar_empresa = round(stats.total_pagar_empresa, 2)
    return resultados, stats


# =============================================================================
# CAMADA DE SAÍDA (geração da planilha formatada — seção 6.2)
# =============================================================================

def montar_aba_detalhamento(wb: Workbook, resultados: list[ResultadoVR]) -> None:
    ws = wb.active
    ws.title = "Detalhamento"

    cabecalhos = [
        "Matrícula", "Nome", "Setor", "Dias a Pagar", "Valor Bruto VR",
        "Desconto Colaborador", "Valor Líquido Empresa", "Elegibilidade",
        "Absenteísmo (%)", "Status",
    ]
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color=COR_HEADER_BG, end_color=COR_HEADER_BG, fill_type="solid")
    header_font = Font(color=COR_HEADER_FONT, bold=True)
    for cel in ws[1]:
        cel.fill = header_fill
        cel.font = header_font
        cel.alignment = Alignment(horizontal="center", vertical="center")

    alerta_fill = PatternFill(start_color=COR_ALERTA_BG, end_color=COR_ALERTA_BG, fill_type="solid")

    for i, r in enumerate(resultados, start=2):
        ws.append([
            r.matricula, r.nome, r.setor, r.dias_a_pagar,
            r.valor_bruto_vr, r.desconto_colaborador, r.valor_liquido_empresa,
            r.elegibilidade, r.absenteismo_pct, r.status,
        ])

        zebra_cor = COR_ZEBRA_PAR if i % 2 == 0 else COR_ZEBRA_IMPAR
        zebra_fill = PatternFill(start_color=zebra_cor, end_color=zebra_cor, fill_type="solid")
        for cel in ws[i]:
            cel.fill = zebra_fill

        for col_letra in ("E", "F", "G"):
            ws[f"{col_letra}{i}"].number_format = FORMATO_MOEDA

        if r.status == "ALERTA":
            ws[f"J{i}"].fill = alerta_fill
            ws[f"J{i}"].font = Font(bold=True)

    # Largura de colunas ajustada automaticamente
    for col_cells in ws.columns:
        comprimento_max = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        letra = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letra].width = comprimento_max + 4

    ws.freeze_panes = "A2"


def montar_aba_resumo(wb: Workbook, stats: Estatisticas) -> None:
    ws = wb.create_sheet("Resumo")

    linhas = [
        ("Total de colaboradores processados", stats.total_colaboradores),
        ("Total a pagar pela empresa", stats.total_pagar_empresa),
        ("Total em alerta de absenteísmo", stats.total_alerta_absenteismo),
        ("Total em período de experiência", stats.total_em_experiencia),
        ("Relatório gerado em", stats.gerado_em.strftime("%d/%m/%Y %H:%M:%S")),
    ]

    header_fill = PatternFill(start_color=COR_HEADER_BG, end_color=COR_HEADER_BG, fill_type="solid")
    header_font = Font(color=COR_HEADER_FONT, bold=True)
    ws.append(["Indicador", "Valor"])
    for cel in ws[1]:
        cel.fill = header_fill
        cel.font = header_font

    for rotulo, valor in linhas:
        ws.append([rotulo, valor])

    ws["B3"].number_format = FORMATO_MOEDA  # total a pagar pela empresa

    for col_cells in ws.columns:
        comprimento_max = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        letra = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letra].width = comprimento_max + 4


def gerar_relatorio_saida(resultados: list[ResultadoVR], stats: Estatisticas, caminho_saida: Path) -> None:
    wb = Workbook()
    montar_aba_detalhamento(wb, resultados)
    montar_aba_resumo(wb, stats)
    try:
        wb.save(caminho_saida)
    except PermissionError as exc:
        raise RoboVRError(
            f"Não foi possível salvar '{caminho_saida.name}'. "
            f"Feche o arquivo se ele já estiver aberto no Excel e tente novamente."
        ) from exc


# =============================================================================
# CLI / ORQUESTRAÇÃO
# =============================================================================

def parse_argumentos(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Robô de cálculo do Vale-Refeição (VR) - Constance Calçados")
    parser.add_argument("--entrada", default="colaboradores.xlsx", help="Nome do arquivo de entrada (.xlsx)")
    parser.add_argument("--saida", default=None, help="Nome do arquivo de saída (.xlsx)")
    parser.add_argument(
        "--data-ref", default=None,
        help="Data de referência para cálculo de elegibilidade/nome do arquivo, formato AAAA-MM-DD. "
             "Padrão: data atual do sistema."
    )
    return parser.parse_args(argv)


def resolver_data_referencia(data_ref_str: str | None) -> date:
    if data_ref_str is None:
        return date.today()
    try:
        return datetime.strptime(data_ref_str, "%Y-%m-%d").date()
    except ValueError:
        raise RoboVRError(f"--data-ref inválida ('{data_ref_str}'). Use o formato AAAA-MM-DD.")


CABECALHO_TERMINAL = "=" * 36


def exibir_cabecalho() -> None:
    print(CABECALHO_TERMINAL)
    print("ROBÔ DE CÁLCULO DE VALE REFEIÇÃO")
    print(CABECALHO_TERMINAL)


def exibir_rodape() -> None:
    print(CABECALHO_TERMINAL)


def exibir_progresso(mensagem: str) -> None:
    print(mensagem)
    time.sleep(INTERVALO_PROGRESSO_SEGUNDOS)


def exibir_resumo(stats: Estatisticas, caminho_saida: Path) -> None:
    print(f"Arquivo: {caminho_saida.name}")
    print(f"Colaboradores processados: {stats.total_colaboradores}")
    print(f"Total a pagar pela empresa: R$ {stats.total_pagar_empresa:,.2f}")
    print(f"Em alerta de absenteísmo: {stats.total_alerta_absenteismo}")
    print(f"Em período de experiência: {stats.total_em_experiencia}")


def aguardar_antes_de_sair() -> None:
    """
    Mantém a janela do terminal aberta ao final da execução.

    Quando o usuário roda o .exe com duplo clique no Windows, a janela do
    console fecha automaticamente assim que o processo termina — sem essa
    pausa, ninguém consegue ler a mensagem de sucesso nem a de erro. Só
    aplica a pausa quando empacotado com PyInstaller (sys.frozen); ao rodar
    via `python robo_vr.py` (desenvolvimento, testes, scripts automatizados),
    a pausa é dispensada para não travar fluxos não interativos.
    """
    if getattr(sys, "frozen", False):
        input("\nPressione ENTER para sair...")


def main(argv: list[str] | None = None) -> int:
    args = parse_argumentos(argv)
    exibir_cabecalho()

    try:
        data_referencia = resolver_data_referencia(args.data_ref)
        nome_saida = args.saida or f"relatorio_vr_{data_referencia.strftime('%m%Y')}.xlsx"
        caminho_saida = caminho_base_execucao() / nome_saida

        exibir_progresso("Lendo arquivo...")
        caminho_entrada = localizar_arquivo_entrada(args.entrada)
        ws = carregar_workbook(caminho_entrada)

        exibir_progresso("Validando colunas...")
        mapa_colunas = validar_estrutura_colunas(ws)
        colaboradores = ler_colaboradores(ws, mapa_colunas)

        exibir_progresso("Aplicando regras...")
        resultados, stats = processar_colaboradores(colaboradores, data_referencia)

        exibir_progresso("Gerando relatório...")
        exibir_progresso("Formatando planilha...")
        gerar_relatorio_saida(resultados, stats, caminho_saida)

        exibir_progresso("Relatório salvo com sucesso.")
        exibir_rodape()
        exibir_resumo(stats, caminho_saida)
        aguardar_antes_de_sair()
        return 0

    except RoboVRError as exc:
        print(f"\n[ERRO] {exc}\n")
        exibir_rodape()
        aguardar_antes_de_sair()
        return 1
    except Exception as exc:  # rede de segurança final: nunca deixar o programa travar sem explicação
        print(f"\n[ERRO INESPERADO] {exc}\n")
        exibir_rodape()
        aguardar_antes_de_sair()
        return 1


if __name__ == "__main__":
    sys.exit(main())